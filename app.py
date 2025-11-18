import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# -------------------------------
# CHANGE THIS TO YOUR ACTUAL FILE PATH
file_path = "Visbreaker_Tag_list_all.xlsx"   # <--- change if your file has a different name
# -------------------------------

# Load the original file
wb = load_workbook(file_path)
ws = wb["Sheet1"]          # or wb.active if the sheet name is different

# Read into pandas (keep_empty_cells is handled automatically with keep_default_na=False)
df = pd.read_excel(file_path, sheet_name="Sheet1", keep_default_na=False)

# Define which columns are "fixed" (they get repeated exactly the same on every new row)
fixed_cols = ["SR. NO.", "P&ID NUMBER", "LINE NUMBER"]

# All the remaining columns are the ones that can contain ALT+ENTER (line breaks)
instrument_cols = [col for col in df.columns if col not in fixed_cols]

# ------------------- EXPAND THE ROWS -------------------
new_rows = []

for _, row in df.iterrows():
    # Convert every instrument column to a list of stripped lines (or [''] if empty)
    col_to_lines = {}
    max_lines = 1

    for col in instrument_cols:
        cell = row[col]
        if cell == "" or (pd.isna(cell)):
            lines = [""]
        else:
            lines = [line.strip() for line in str(cell).split("\n") if line.strip() != ""]
            if len(lines) == 0:
                lines = [""]
        col_to_lines[col] = lines
        max_lines = max(max_lines, len(lines))  # <-- FIXED: Added missing )

    # Pad shorter columns with empty strings so every column has the same number of lines
    for col in instrument_cols:
        if len(col_to_lines[col]) < max_lines:
            col_to_lines[col] += [""] * (max_lines - len(col_to_lines[col]))

    # Fixed values for this original row
    sr_no   = row["SR. NO."] if pd.notna(row["SR. NO."]) else ""
    pid_no  = row["P&ID NUMBER"] if pd.notna(row["P&ID NUMBER"]) else ""
    line_no = row["LINE NUMBER"] if pd.notna(row["LINE NUMBER"]) else ""

    # Create the new sub-rows
    for i in range(max_lines):
        new_row = {
            "SR. NO.": sr_no,
            "P&ID NUMBER": pid_no,
            "LINE NUMBER": line_no,
        }
        for col in instrument_cols:
            new_row[col] = col_to_lines[col][i]
        new_rows.append(new_row)

# Create the new DataFrame (keeps the original column order)
new_df = pd.DataFrame(new_rows, columns=df.columns)

# ------------------- WRITE BACK TO EXCEL WITH MERGES -------------------
# We will overwrite the sheet (or you can create a new one with wb.create_sheet("Fixed"))
ws.delete_rows(1, ws.max_row + 10)   # clear everything

# Write header
for col_idx, header in enumerate(new_df.columns, 1):
    ws.cell(row=1, column=col_idx, value=header)

# Write data
for r_idx, row in enumerate(dataframe_to_rows(new_df, index=False, header=False), 2):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# ------------------- MERGE P&ID NUMBER and LINE NUMBER for identical consecutive groups -------------------
center_alignment = Alignment(horizontal="center", vertical="center")  # <-- MOVED UP: Define before using

start_row = 2
prev_pid = ws.cell(2, 2).value or ""
prev_line = ws.cell(2, 3).value or ""

for current_row in range(3, ws.max_row + 2):   # +2 so the last group gets processed
    curr_pid  = ws.cell(current_row, 2).value or ""
    curr_line = ws.cell(current_row, 3).value or ""

    if curr_pid != prev_pid or curr_line != prev_line:
        # end previous group
        if current_row - start_row > 1:   # more than one row → merge
            ws.merge_cells(start_row=start_row, start_column=2, end_row=current_row-1, end_column=2)
            ws.merge_cells(start_row=start_row, start_column=3, end_row=current_row-1, end_column=3)
            for r in range(start_row, current_row):
                ws.cell(r, 2).alignment = center_alignment
                ws.cell(r, 3).alignment = center_alignment

        start_row = current_row
        prev_pid  = curr_pid
        prev_line = curr_line

# Merge the very last group (the loop above misses it)
if ws.max_row - start_row > 0:  # at least 2 rows in last group
    ws.merge_cells(start_row=start_row, start_column=2, end_row=ws.max_row, end_column=2)
    ws.merge_cells(start_row=start_row, start_column=3, end_row=ws.max_row, end_column=3)
    for r in range(start_row, ws.max_row + 1):
        ws.cell(r, 2).alignment = center_alignment
        ws.cell(r, 3).alignment = center_alignment

# Save the fixed file
wb.save("Visbreaker_Tag_list_all.xlsx")

print("Done! Fixed file created → Visbreaker_Tag_list_all_FIXED.xlsx")